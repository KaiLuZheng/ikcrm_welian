
#!/usr/bin/python3
#-*- coding=utf8 -*-

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
import re
import os, sys

'''
CORE_DIR = os.path.dirname(os.path.realpath(__file__))
FOO_DIR = os.path.dirname(CORE_DIR)
sys.path.append(FOO_DIR)
'''


#####################
# for the openpyxl of sheet used
#####################
FIRST_ROW = 1
FIRST_COL = 'A'
JUST_ONE_SHEET = True


class dealExcelCore(object):
    def __init__(self, filename = 'outExcel.xlsx'):
        self.wb = Workbook()
        self.ws = self.wb.active

        self.filename = filename

    def setlabels(self, labels):
        self.setDate(row = 1, infos = labels)

    def setDate(self, row, infos): # infos is a list
        sheet = self.ws 
        for i in range(0, len(infos)):
            sheet.cell(row = row, column = 1+i).value = infos[i]

    def saveExcel(self):
        self.wb.save(self.filename)


    def cells2list(self, cells): # redefine the list of cells
        # try the type of cell
        if isinstance(cells[0], type('string')):
            return cells
        elif isinstance(cells[0], type([])):
            return cells
        return [cell.value for cell in cells]



if __name__ == '__main__':
    t1 = ['a', 'b', 'c']
    t2 = ['1','2','3']
    a = dealExcelCore('exceltest.xlsx')
    a.setlabels(t1)
    a.setDate(2,t2)
    a.saveExcel()

